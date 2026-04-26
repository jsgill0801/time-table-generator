"""Analyze the reference Excel file - deep dive into structure."""
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

wb = load_workbook(r"Lecture_Time_Table_Win'26_v6.xlsx")
ws = wb.active

# Check merged cells in rows 1-6 
print("=== MERGED CELLS ===")
for mc in sorted(ws.merged_cells.ranges, key=lambda x: (x.min_row, x.min_col)):
    if mc.min_row <= 10:
        print(f"  {mc}")

print()

# Row 1 and 2 - title
print("=== TITLE ROWS ===")
for r in range(1, 4):
    for c in range(1, 40):
        v = ws.cell(r, c).value
        if v is not None:
            print(f"  {get_column_letter(c)}{r} = '{v}'")

# Day header structure
print()
print("=== DAY HEADERS (Row 4) ===")
for c in range(1, 40):
    v = ws.cell(4, c).value
    if v is not None:
        print(f"  Col {c} ({get_column_letter(c)}): '{v}'")

# Slot/sub-header columns within each day
print()
print("=== SLOT ROW (Row 5) ===")
for c in range(1, 40):
    v = ws.cell(5, c).value
    if v is not None:
        print(f"  Col {c} ({get_column_letter(c)}): '{v}'")

# Row 6 - the first time slot header and slot names
print()
print("=== ROW 6 (first time slot) ===")
for c in range(1, 40):
    v = ws.cell(6, c).value
    if v is not None:
        print(f"  Col {c} ({get_column_letter(c)}): '{v}'")

# Check what the sub-columns are within each day
# E.g. Monday starts at col D (4)
# What are cols D, E, F, G, H, I?
print()
print("=== COLUMN MEANING (from Row 7 data) ===")
print("  D=course_code, E=course_name, F=ltpc, G=category, H=faculty_code, I=room")
print("  Col C is a separator")
print("  Day columns: D-I (Mon=6cols), K-P (Tue=6cols), R-W (Wed=6cols), Y-AD (Thu=6cols), AF-AK (Fri=6cols)")

# Verify the column pattern for each day
print()
print("=== Day column starts (1-indexed) ===")
print("  Monday: D = col 4")
print("  Tuesday: K = col 11")
print("  Wednesday: R = col 18")
print("  Thursday: Y = col 25")
print("  Friday: AF = col 32")

# Check how many rows are in each time slot group
print()
print("=== TIME SLOT GROUPS ===")
for r in range(6, 200):
    v = ws.cell(r, 1).value  # column A = Time
    b = ws.cell(r, 2).value  # column B = Batch
    if v is not None:
        print(f"  Row {r}: TIME='{v}'")
    elif b is not None:
        print(f"  Row {r}: BATCH='{b}'")
    else:
        # Check if there's any data in this row
        has_data = any(ws.cell(r, c).value for c in range(1, 40))
        if has_data:
            first_val = None
            for c in range(1, 40):
                if ws.cell(r, c).value:
                    first_val = f"{get_column_letter(c)}={ws.cell(r, c).value}"
                    break
            print(f"  Row {r}: (no time/batch) first={first_val}")

    if r > 150 and not any(ws.cell(r, c).value for c in range(1, 40)):
        # Stop after 5 empty rows
        empty_count = sum(1 for rr in range(r, r+5) if not any(ws.cell(rr, c).value for c in range(1, 40)))
        if empty_count >= 5:
            break

# Check row colors/fills for time slot rows
print()
print("=== ROW COLORS (sample) ===")
for r in [6, 7, 8, 9, 10, 11]:
    cell_b = ws.cell(r, 2)
    fill = cell_b.fill
    if fill and fill.start_color:
        try:
            print(f"  Row {r}, Col B: fill={fill.start_color.rgb}, type={fill.patternType}")
        except:
            print(f"  Row {r}, Col B: fill type={fill.patternType}")
