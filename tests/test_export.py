"""
Tests for the ExportService.

Verifies that the Excel workbook is generated correctly
with proper structure for the four supported exports:
overall (single sheet), faculty-wise, batch-wise, and room-wise.
"""

import os
import uuid
import pytest

from backend.services.export_service import ExportService


@pytest.fixture
def export_tmp_dir():
    """Use a plain temp directory because pytest's tmp_path ACL is restricted here."""
    root = os.path.abspath(".test-output")
    os.makedirs(root, exist_ok=True)
    path = os.path.join(root, f"ttg-export-{uuid.uuid4().hex}")
    os.makedirs(path, exist_ok=True)
    return path


# -----------------------------------------------------------------
#  Sample timetable data for testing
# -----------------------------------------------------------------

def _sample_timetable():
    """Return a minimal timetable dataset for export testing."""
    return [
        {
            "auto_id": 1,
            "batch_course_id": 1,
            "faculty_code": "PD",
            "classroom_name": "LT-1",
            "slot_id": "MON-0800",
            "day_of_week": "Monday",
            "start_time": "08:00",
            "end_time": "08:50",
            "slot_name": "Slot-1",
            "course_code": "IT205",
            "course_name": "Data Structures",
            "ltpc": "3-0-0-3",
            "category_name": "Core",
            "batch_label": "BTech Sem-IV (ICT) Sec-A",
        },
        {
            "auto_id": 2,
            "batch_course_id": 2,
            "faculty_code": "AV",
            "classroom_name": "CEP211",
            "slot_id": "TUE-0900",
            "day_of_week": "Tuesday",
            "start_time": "09:00",
            "end_time": "09:50",
            "slot_name": "Slot-7",
            "course_code": "MA206",
            "course_name": "Probability",
            "ltpc": "3-1-0-4",
            "category_name": "Core",
            "batch_label": "BTech Sem-IV (ICT) Sec-A",
        },
        {
            "auto_id": 3,
            "batch_course_id": 3,
            "faculty_code": "PD",
            "classroom_name": "LT-1",
            "slot_id": "MON-0800",
            "day_of_week": "Monday",
            "start_time": "08:00",
            "end_time": "08:50",
            "slot_name": "Slot-1",
            "course_code": "HM108",
            "course_name": "English Communication",
            "ltpc": "2-0-0-2",
            "category_name": "Humanities",
            "batch_label": "BTech Sem-IV (CS)",
        },
    ]


# =================================================================
#  Tests
# =================================================================

class TestExportService:
    """Tests for the Excel export service."""

    def test_generates_xlsx_file(self, export_tmp_dir, monkeypatch):
        """The overall export should create an .xlsx file on disk."""
        # Override the output directory to use a temp path
        monkeypatch.setattr(
            "backend.services.export_service.OUTPUT_DIR",
            str(export_tmp_dir),
        )

        service = ExportService(_sample_timetable())
        filepath = service.generate_overall()

        assert os.path.exists(filepath)
        assert filepath.endswith(".xlsx")
        assert os.path.getsize(filepath) > 0

    def test_workbook_has_correct_sheets(self, export_tmp_dir, monkeypatch):
        """The overall workbook should have a single sheet."""
        monkeypatch.setattr(
            "backend.services.export_service.OUTPUT_DIR",
            str(export_tmp_dir),
        )

        service = ExportService(_sample_timetable())
        filepath = service.generate_overall()

        from openpyxl import load_workbook
        wb = load_workbook(filepath)

        sheet_names = wb.sheetnames

        assert "Time-Table" in sheet_names
        assert len(sheet_names) == 1

    def test_master_sheet_has_title(self, export_tmp_dir, monkeypatch):
        """The main sheet should contain the grouped timetable title."""
        monkeypatch.setattr(
            "backend.services.export_service.OUTPUT_DIR",
            str(export_tmp_dir),
        )

        service = ExportService(_sample_timetable())
        filepath = service.generate_overall()

        from openpyxl import load_workbook
        wb = load_workbook(filepath)
        ws = wb["Time-Table"]

        assert ws.cell(row=2, column=1).value == "Lecture Time-Table"

    def test_faculty_sheet_has_data(self, export_tmp_dir, monkeypatch):
        """Faculty-wise export should create one sheet per faculty."""
        monkeypatch.setattr(
            "backend.services.export_service.OUTPUT_DIR",
            str(export_tmp_dir),
        )

        service = ExportService(_sample_timetable())
        filepath = service.generate_faculty_wise()

        from openpyxl import load_workbook
        wb = load_workbook(filepath)
        assert "PD" in wb.sheetnames
        assert "AV" in wb.sheetnames

    def test_empty_timetable(self, export_tmp_dir, monkeypatch):
        """An empty timetable should still produce a valid file."""
        monkeypatch.setattr(
            "backend.services.export_service.OUTPUT_DIR",
            str(export_tmp_dir),
        )

        service = ExportService([])
        filepath = service.generate_overall()

        assert os.path.exists(filepath)
        assert filepath.endswith(".xlsx")
