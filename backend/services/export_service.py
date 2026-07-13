"""
Excel export service.

Generates four separate Excel workbooks that replicate the exact layout of
the university's official "Lecture_Time_Table" file:

    1. Overall timetable   — single sheet, all batches grouped by time-slot
    2. Faculty-wise         — one sheet per faculty, same grid format
    3. Batch-wise           — one sheet per batch, same grid format
    4. Room-wise            — one sheet per room, same grid format

Column layout per day (6 columns):
    course_code | course_name | ltpc | category_name | faculty_code | classroom_name

Days start at columns:  D(4)  K(11)  R(18)  Y(25)  AF(32)
Column C is a visual separator.
"""

import os
import logging
import json
from collections import defaultdict
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)
DEBUG_LOG_PATH = os.path.join(
    os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))),
    "debug-ecec21.log",
)

OUTPUT_DIR = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    "output",
)

WEEKDAYS = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]

# Column where each day block begins (1-indexed)
DAY_COL_START = {
    "Monday": 4,     # D
    "Tuesday": 11,   # K
    "Wednesday": 18, # R
    "Thursday": 25,  # Y
    "Friday": 32,    # AF
}
# Each day block spans 6 columns: code, name, ltpc, category, faculty, room
DAY_SPAN = 6

# Detail fields within each day block (in order)
DETAIL_KEYS = [
    "course_code",
    "course_name",
    "ltpc",
    "category_name",
    "faculty_code",
    "classroom_name",
]

# ---------- colour palette (matches the reference file) ----------
TITLE_FONT = Font(name="Calibri", size=14, bold=True, color="C00000")
SUBTITLE_FONT = Font(name="Calibri", size=12, bold=True, color="C00000")

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

SLOT_ROW_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
SLOT_ROW_FONT = Font(name="Calibri", size=10, bold=True, color="1F4E79")

TIME_FONT = Font(name="Calibri", size=10, bold=True, color="1F4E79")
BATCH_FONT = Font(name="Calibri", size=10, bold=True, color="333333")
CELL_FONT = Font(name="Calibri", size=10)

# Alternating batch row fills for readability
BATCH_FILL_LIGHT = PatternFill(start_color="EBF1DE", end_color="EBF1DE", fill_type="solid")
BATCH_FILL_ALT   = PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid")
SEPARATOR_FILL   = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style="thin", color="B4C6E7"),
    right=Side(style="thin", color="B4C6E7"),
    top=Side(style="thin", color="B4C6E7"),
    bottom=Side(style="thin", color="B4C6E7"),
)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


class ExportService:
    """Build Excel workbooks from timetable rows and slot definitions."""

    def __init__(self, timetable_rows: list[dict], slots: list[dict] | None = None):
        self.rows = timetable_rows or []
        self.slots = slots or []
        self._run_id = f"export_{int(datetime.now().timestamp() * 1000)}"

        # Pre-index data
        self.by_batch: dict[str, list[dict]] = defaultdict(list)
        self.by_faculty: dict[str, list[dict]] = defaultdict(list)
        self.by_room: dict[str, list[dict]] = defaultdict(list)

        for row in self.rows:
            self.by_batch[row["batch_label"]].append(row)
            if row.get("faculty_code"):
                self.by_faculty[row["faculty_code"]].append(row)
            if row.get("classroom_name"):
                self.by_room[row["classroom_name"]].append(row)

        self.time_ranges = self._extract_time_ranges()
        self.slot_grid = self._build_slot_grid()
        self._debug_log(
            hypothesis_id="H1",
            location="export_service.py:__init__",
            message="Initialized export service",
            data={
                "rows_count": len(self.rows),
                "slots_count": len(self.slots),
                "time_ranges_count": len(self.time_ranges),
                "sample_days": sorted(list({str(r.get("day_of_week", "")) for r in self.rows}))[:10],
            },
        )

        os.makedirs(OUTPUT_DIR, exist_ok=True)

    # ------------------------------------------------------------------
    #  Public API — each method returns a file path
    # ------------------------------------------------------------------

    def generate_overall(self) -> str:
        """Overall timetable — single sheet, every batch."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Time-Table"
        self._write_overall_sheet(ws, self.rows)
        return self._save(wb, "Overall_Timetable")

    def generate_faculty_wise(self) -> str:
        """Faculty-wise — one sheet per faculty."""
        wb = Workbook()
        wb.remove(wb.active)
        for fcode in sorted(self.by_faculty.keys()):
            ws = wb.create_sheet(self._safe_sheet_name(fcode, wb))
            self._write_overall_sheet(ws, self.by_faculty[fcode], subtitle=f"Faculty: {fcode}")
        if not wb.sheetnames:
            ws = wb.create_sheet("Empty")
            ws["A1"] = "No faculty data available."
        return self._save(wb, "Faculty_Wise_Timetable")

    def generate_batch_wise(self) -> str:
        """Batch-wise — one sheet per batch."""
        wb = Workbook()
        wb.remove(wb.active)
        for blabel in sorted(self.by_batch.keys()):
            ws = wb.create_sheet(self._safe_sheet_name(blabel, wb))
            self._write_overall_sheet(ws, self.by_batch[blabel], subtitle=f"Batch: {blabel}")
        if not wb.sheetnames:
            ws = wb.create_sheet("Empty")
            ws["A1"] = "No batch data available."
        return self._save(wb, "Batch_Wise_Timetable")

    def generate_room_wise(self) -> str:
        """Room-wise — one sheet per room."""
        wb = Workbook()
        wb.remove(wb.active)
        for rname in sorted(self.by_room.keys()):
            ws = wb.create_sheet(self._safe_sheet_name(rname, wb))
            self._write_overall_sheet(ws, self.by_room[rname], subtitle=f"Room: {rname}")
        if not wb.sheetnames:
            ws = wb.create_sheet("Empty")
            ws["A1"] = "No room data available."
        return self._save(wb, "Room_Wise_Timetable")

    # ------------------------------------------------------------------
    #  Sheet writer — reproduces the exact reference layout
    # ------------------------------------------------------------------

    def _write_overall_sheet(self, ws, rows: list[dict], subtitle: str | None = None):
        """
        Writes the university-format timetable into the given worksheet.

        Layout:
            Row 1:  merged title
            Row 2:  merged subtitle
            Row 3:  (blank)
            Row 4:  header row — Time | Batch | (sep) | Monday(merged 6) | … | Friday(merged 6)
            Row 5:  (sub-header, blank in data — used for column C separator marker)
            For each time-range:
                Slot row:  time in A, slot names merged across each day block
                Data rows: one per batch, columns filled per-day
        """
        last_col = 37  # AK

        # ---- Row 1: University title ----
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
        ws["A1"] = "Dhirubhai Ambani University (School of Technology)"
        ws["A1"].font = TITLE_FONT
        ws["A1"].alignment = CENTER

        # ---- Row 2: Subtitle ----
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_col)
        ws["A2"] = subtitle or "Lecture Time-Table"
        ws["A2"].font = SUBTITLE_FONT
        ws["A2"].alignment = CENTER

        # ---- Row 3: blank spacer ----
        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=last_col)

        # ---- Row 4: Day headers ----
        hdr_row = 4
        for label, col in [("Time", 1), ("Batch", 2)]:
            cell = ws.cell(hdr_row, col, label)
            self._style_header(cell)

        for day, start_col in DAY_COL_START.items():
            end_col = start_col + DAY_SPAN - 1
            ws.merge_cells(
                start_row=hdr_row, start_column=start_col,
                end_row=hdr_row, end_column=end_col,
            )
            cell = ws.cell(hdr_row, start_col, day)
            self._style_header(cell)

        # ---- Build the batch/time grouping ----
        # Group rows by time-range, then by batch within each time-range
        time_batch_map: dict[str, dict[str, list[dict]]] = {}
        for tr in self.time_ranges:
            time_batch_map[tr] = defaultdict(list)

        missing_time_ranges = 0
        for row in rows:
            tr = self._time_range_label(row["start_time"], row["end_time"])
            if tr and tr in time_batch_map:
                time_batch_map[tr][row["batch_label"]].append(row)
            elif tr:
                missing_time_ranges += 1

        self._debug_log(
            hypothesis_id="H3",
            location="export_service.py:_write_overall_sheet",
            message="Built time-batch map",
            data={
                "sheet_title": ws.title,
                "input_rows": len(rows),
                "available_time_ranges": len(self.time_ranges),
                "missing_time_ranges": missing_time_ranges,
            },
        )

        # ---- Write data rows ----
        current_row = 6  # leave row 5 blank (matches reference)

        for time_range in self.time_ranges:
            batch_groups = time_batch_map.get(time_range, {})
            batches_sorted = sorted(batch_groups.keys())

            if not batches_sorted:
                # Still write the slot row with no batches
                batches_sorted = []

            # --- Slot header row ---
            slot_row = current_row

            # Time cell (will be merged down later)
            time_cell = ws.cell(slot_row, 1, time_range)
            time_cell.font = TIME_FONT
            time_cell.alignment = CENTER
            time_cell.fill = SLOT_ROW_FILL
            time_cell.border = THIN_BORDER

            # Slot names per day (merged across day columns)
            for day, start_col in DAY_COL_START.items():
                end_col = start_col + DAY_SPAN - 1
                ws.merge_cells(
                    start_row=slot_row, start_column=start_col,
                    end_row=slot_row, end_column=end_col,
                )
                slot_name = self.slot_grid.get(time_range, {}).get(day, "")
                cell = ws.cell(slot_row, start_col, slot_name)
                cell.font = SLOT_ROW_FONT
                cell.alignment = CENTER
                cell.fill = SLOT_ROW_FILL
                cell.border = THIN_BORDER

            current_row += 1  # move past slot header

            if not batches_sorted:
                current_row += 1
                continue

            # --- Data rows (one row per batch) ---
            # For batches that have entries across multiple days,
            # some may have more than one entry per day (e.g. electives).
            # In the reference, extra rows simply continue below the batch row
            # without repeating the batch label.

            for b_idx, batch_label in enumerate(batches_sorted):
                day_entries = batch_groups[batch_label]

                # Group by day, preserving order
                by_day: dict[str, list[dict]] = defaultdict(list)
                for entry in day_entries:
                    by_day[entry["day_of_week"]].append(entry)

                unmapped_days = sorted([d for d in by_day.keys() if d not in WEEKDAYS])
                if unmapped_days:
                    self._debug_log(
                        hypothesis_id="H2",
                        location="export_service.py:_write_overall_sheet",
                        message="Found entries with unmapped day values",
                        data={
                            "sheet_title": ws.title,
                            "time_range": time_range,
                            "batch_label": batch_label,
                            "unmapped_days": unmapped_days[:10],
                        },
                    )

                # Max rows needed for this batch
                max_sub_rows = max((len(v) for v in by_day.values()), default=1) if by_day else 1

                for sub_row_idx in range(max_sub_rows):
                    row_num = current_row + sub_row_idx
                    row_fill = BATCH_FILL_LIGHT if b_idx % 2 == 0 else BATCH_FILL_ALT

                    # Batch label (only on the first sub-row)
                    if sub_row_idx == 0:
                        bcell = ws.cell(row_num, 2, batch_label)
                        bcell.font = BATCH_FONT
                        bcell.alignment = LEFT
                        bcell.fill = row_fill
                        bcell.border = THIN_BORDER

                        # Merge batch cell down if multiple sub-rows
                        if max_sub_rows > 1:
                            ws.merge_cells(
                                start_row=row_num, start_column=2,
                                end_row=row_num + max_sub_rows - 1, end_column=2,
                            )

                    # Separator column C
                    sep = ws.cell(row_num, 3, "")
                    sep.fill = SEPARATOR_FILL

                    # Day blocks
                    for day, start_col in DAY_COL_START.items():
                        entries_for_day = by_day.get(day, [])
                        entry = entries_for_day[sub_row_idx] if sub_row_idx < len(entries_for_day) else None

                        for offset, key in enumerate(DETAIL_KEYS):
                            col = start_col + offset
                            value = (entry or {}).get(key, "")
                            cell = ws.cell(row_num, col, value)
                            cell.font = CELL_FONT
                            cell.alignment = LEFT if offset == 1 else CENTER
                            cell.fill = row_fill
                            cell.border = THIN_BORDER

                current_row += max_sub_rows

            # Merge the time cell down to cover all batch rows
            time_end_row = current_row - 1
            if time_end_row > slot_row:
                ws.merge_cells(
                    start_row=slot_row, start_column=1,
                    end_row=time_end_row, end_column=1,
                )

            current_row += 1  # blank separator row between time blocks

        # ---- Column widths (match reference) ----
        self._set_column_widths(ws)

    # ------------------------------------------------------------------
    #  Styling helpers
    # ------------------------------------------------------------------

    @staticmethod
    def _style_header(cell):
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    @staticmethod
    def _set_column_widths(ws):
        widths = {
            "A": 16, "B": 28, "C": 3,
            # Monday D-I
            "D": 12, "E": 40, "F": 14, "G": 22, "H": 10, "I": 14,
            # separator
            "J": 3,
            # Tuesday K-P
            "K": 12, "L": 40, "M": 14, "N": 22, "O": 10, "P": 14,
            # separator
            "Q": 3,
            # Wednesday R-W
            "R": 12, "S": 40, "T": 14, "U": 22, "V": 10, "W": 14,
            # separator
            "X": 3,
            # Thursday Y-AD
            "Y": 12, "Z": 40, "AA": 14, "AB": 22, "AC": 10, "AD": 14,
            # separator
            "AE": 3,
            # Friday AF-AK
            "AF": 12, "AG": 40, "AH": 14, "AI": 22, "AJ": 10, "AK": 14,
        }
        for col_letter, w in widths.items():
            ws.column_dimensions[col_letter].width = w

    # ------------------------------------------------------------------
    #  Data helpers
    # ------------------------------------------------------------------

    def _extract_time_ranges(self) -> list[str]:
        """Return unique time-range labels sorted chronologically."""
        ranges: dict[str, tuple[str, str]] = {}
        source = self.slots if self.slots else self.rows
        for item in source:
            label = self._time_range_label(item.get("start_time"), item.get("end_time"))
            if label:
                ranges[label] = (
                    self._normalize_time(item.get("start_time")),
                    self._normalize_time(item.get("end_time")),
                )
        result = [lbl for lbl, _ in sorted(ranges.items(), key=lambda x: x[1])]
        self._debug_log(
            hypothesis_id="H4",
            location="export_service.py:_extract_time_ranges",
            message="Extracted time ranges",
            data={
                "source": "slots" if self.slots else "rows",
                "time_ranges_count": len(result),
                "first_time_ranges": result[:5],
            },
        )
        return result

    def _build_slot_grid(self) -> dict[str, dict[str, str]]:
        """Map (time_range, day) -> slot_name."""
        grid: dict[str, dict[str, str]] = defaultdict(dict)
        for slot in self.slots:
            tr = self._time_range_label(slot.get("start_time"), slot.get("end_time"))
            day = slot.get("day_of_week")
            if tr and day in WEEKDAYS:
                grid[tr][day] = slot.get("slot_name", "")
        self._debug_log(
            hypothesis_id="H5",
            location="export_service.py:_build_slot_grid",
            message="Built slot grid",
            data={
                "slot_grid_time_ranges": len(grid.keys()),
                "sample_keys": list(grid.keys())[:5],
            },
        )
        return grid

    @staticmethod
    def _time_range_label(start_time, end_time) -> str:
        start = ExportService._normalize_time(start_time)
        end = ExportService._normalize_time(end_time)
        if not start or not end:
            return ""
        return f"{start} - {end}"

    @staticmethod
    def _normalize_time(value) -> str:
        parts = str(value or "").split(":")
        if len(parts) < 2:
            return ""
        return parts[0].zfill(2) + ":" + parts[1].zfill(2)

    # ------------------------------------------------------------------
    #  File I/O helpers
    # ------------------------------------------------------------------

    def _save(self, wb: Workbook, prefix: str) -> str:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{prefix}_{timestamp}.xlsx"
        filepath = os.path.join(OUTPUT_DIR, filename)
        wb.save(filepath)
        logger.info("Timetable exported to: %s", filepath)
        return filepath

    @staticmethod
    def _safe_sheet_name(raw: str, wb: Workbook) -> str:
        invalid = '[]:*?/\\'
        clean = "".join("-" if ch in invalid else ch for ch in str(raw))
        clean = clean.strip() or "Sheet"
        base = clean[:31]
        name = base
        idx = 2
        while name in wb.sheetnames:
            suffix = f" ({idx})"
            name = base[: 31 - len(suffix)] + suffix
            idx += 1
        return name

    def _debug_log(self, hypothesis_id: str, location: str, message: str, data: dict):
        pass
